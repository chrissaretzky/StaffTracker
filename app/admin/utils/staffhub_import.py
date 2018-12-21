from openpyxl import load_workbook
from openpyxl import worksheet
import re
import calendar
from app.admin.utils.data_import import Excel_Import
from datetime import datetime, timedelta


class ValidationError(Exception):
    pass


class Staffhub_Import(Excel_Import):
    valid_colors = {
        "blank": "00000000",
        "white": "FFFFFFFF",
        "grey": "FFE0E0E0",
        "light-yellow": "FFFFFF99",
        "dark-yellow": "FFF9D748",
        "dark-yellow2": "FFFFB902",
        "light-green": "FFCCFFCC",
        "dark-green": "FF99CC66",
        "light-pink": "FFFFCCCC",
        "dark-pink": "FFFF6699",
        "light-blue": "FF99CCFF",
        "dark-blue": "FF6699CC",
        "light-purple": "FFCCCCFF",
        "dark-purple": "FF9966CC"
    }
    valid_regex = {
        "header-date":
        "(?:Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday|Tues|Thur|Thurs|Sun|Mon|Tue|Wed|Thu|Fri|Sat),\\s+(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Sept|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\\s+\\d+",
        "body-time": [
            "\\d+\\s+[A-Z]M\\s+-\\s+\\d+\\s+[A-Z]M",
            "\\d+:\\d+\\s+[A-Z]M\\s+-\\s+\\d+:\\d+\\s+[A-Z]M",
            "\\d+\\s+[A-Z]M\\s+-\\s+\\d+:\\d+\\s+[A-Z]M",
            "\\d+:\\d+\\s+[A-Z]M\\s+-\\s+\\d+\\s+[A-Z]M",
        ]
    }
    valid_time_off = ["Vacation", "Sick", "FRL", "Stat", "BDay", "Off"]
    time_off_request = {"color": "FFF9D748", "text": "Time off request"}

    body_row_offset = 4
    body_col_offset = 4
    body_col_offsetV = 'D'
    problem_cells = {}

    def __init__(self, excel_file, yearof):
        super().__init__()
        self.yearof = yearof
        self.sh = load_workbook(excel_file).active
        self.isValidated = self.run_validation()
        if not self.isValidated:
            print(self.log)
            raise ValidationError(
                'Sheet is not is not in the correct format please see log file'
            )

    def generate_employees_dict(self, sh):
        emps = {}

        for i in range(self.body_row_offset, sh.max_row):
            name = sh.cell(i, 1)
            if len(emps) == 0:
                emps[i] = name.value
            else:
                emps[name.row] = name.value
        self.log['employees'] = list(emps.values())
        return emps

    def generate_dates_dict(self, sh):
        dates = {}
        max_ = sh.max_column + 1
        for i in range(4, max_):
            day = sh.cell(1, i)
            dates[day.column] = self.create_date(day.value, self.yearof)
        self.log['dates'] = list(dates.values())
        return dates

    def create_date(self, day, year):
        month = day.split(' ')[1].strip()
        daynum = int(day.split(' ')[2].strip())
        month = self.change_cal_name_into_num(month, calendar.month_abbr)

        return datetime(int(year), month, daynum)

    def change_cal_name_into_num(self, cal_name, cal_names):
        cal_num = None

        for i in range(0, len(cal_names)):
            if cal_name == cal_names[i]:
                cal_num = i
        return cal_num

    def run_validation(self):
        c = self.check_sheet_for_invalid_colors()
        t = self.check_sheet_body_for_valid_times()
        d = self.check_sheet_header_for_valid_dates()
        if c and t and d:
            return True
        else:
            return False

    def check_sheet_for_invalid_colors(self):
        problem_cells = []

        for i in range(1, self.sh.max_column):
            for ii in range(1, self.sh.max_row):
                cell = self.sh.cell(ii, i)
                color = cell.fill.start_color.rgb
                if color not in self.valid_colors.values():
                    print(color)
                    problem_cells.append(cell.coordinate)
        if len(problem_cells) > 0:
            self.log['invalid_colors'] = problem_cells
            return False
        else:
            self.log['invalid_colors'] = None
            return True

    def check_sheet_header_for_valid_dates(self):
        problem_cells = []
        for i in range(self.body_row_offset, self.sh.max_column):
            cell = self.sh.cell(1, i)
            value = cell.value
            m = re.match(self.valid_regex['header-date'], value)
            if not m:
                problem_cells.append(cell.coordinate)
            elif m.span()[1] < len(value):
                problem_cells.append(cell.coordinate)

        if len(problem_cells) > 0:
            self.log['invalid_header'] = problem_cells
            return False
        else:
            self.log['invalid_header'] = None
            return True

    def check_sheet_body_for_valid_times(self):
        problem_cells = []
        max_ = self.sh.max_column + 1

        for i in range(self.body_col_offset, max_):
            for ii in range(self.body_row_offset, self.sh.max_row):
                cell = self.sh.cell(ii, i)
                value = cell.value
                color = cell.fill.start_color.rgb
                if color != self.valid_colors['dark-yellow2']:
                    if value:
                        value = value.split('\n')[0]
                    if value:
                        matches = []
                        for reg in self.valid_regex['body-time']:
                            m = re.match(reg, value)
                            matches.append(m)
                        if value not in self.valid_time_off:
                            if all(match is None for match in matches):
                                problem_cells.append(cell.coordinate)
        if len(problem_cells) > 0:
            self.log['invalid_times'] = problem_cells
            return False
        else:
            self.log['invalid_times'] = None
            return True


class Shifts_Import(Staffhub_Import):
    shifts_schema = {
        'first_name': {
            'type': 'string'
        },
        'last_name': {
            'type': 'string'
        },
        'shiftstart': {
            'type': 'datetime'
        },
        'shiftend': {
            'type': 'datetime'
        },
        'isic': {
            'type': 'number',
            'allowed': [0, 1]
        },
        'training': {
            'type': 'number',
            'allowed': [0, 1]
        },
        'otbanked': {
            'type': 'number',
            'allowed': [0, 1]
        },
        'otpaidout': {
            'type': 'number',
            'allowed': [0, 1]
        },
        'comment': {
            'type': 'string',
            'nullable': True
        }
    }

    def __init__(self, excel_file, yearof):
        super().__init__(excel_file, yearof)
        self.data = self.generate_shifts()
        self.valid_data = self.validate_import(self.data, self.shifts_schema)

    def generate_shifts(self):

        shifts = []
        self.log['records'] = 0

        dates = self.generate_dates_dict(self.sh)
        emps = self.generate_employees_dict(self.sh)
        max_ = self.sh.max_column + 1

        for i in range(self.body_col_offset, max_):
            for ii in range(self.body_row_offset, self.sh.max_row):
                shift = self.create_shift_data(
                    self.sh.cell(ii, i), dates, emps)
                if shift:
                    shifts.append(shift)
                    self.log['records'] = 1 + self.log['records']
                    print('added shift: ' + str(self.log['records']))
        self.log['type'] = 'shifts'
        print('shifts generated: ' + str(self.log))
        return shifts

    def create_shift_data(self, cell, dates, emps):
        color = cell.fill.start_color.rgb
        IsIC = 0
        OTBanked = 0
        OTPaidOut = 0
        training = 0

        if cell.value:
            c = cell.value.split('\n')
            value = cell.value
            comment = None
            if len(c) > 1:
                value = c[0]
                comment = str(c[1])

            matches = []
            for reg in self.valid_regex['body-time']:
                m = re.match(reg, value)
                matches.append(m)

            if not all(match is None for match in matches):
                start_end = self.create_start_end_time(value)
                if self.sh.cell(cell.row, 2).value == 'Training':
                    training = 1
                if self.sh.cell(cell.row, 2).value == 'Agents':
                    vc = self.valid_colors

                    if color == vc['light-yellow']:
                        OTBanked = 1
                    if color == vc['dark-purple']:
                        training = 1
                    if color == vc['dark-yellow']:
                        OTPaidOut = 1
                    if color == vc['dark-blue'] or color == vc[
                            'dark-green'] or color == vc['dark-pink']:
                        IsIC = 1

                if int(start_end[0].split(':')[0]) > int(
                        start_end[1].split(':')[0]):
                    start_end[1] = str(dates[cell.column].date() +
                                       timedelta(days=1)) + ' ' + start_end[1]
                else:
                    start_end[1] = str(
                        dates[cell.column].date()) + ' ' + start_end[1]

                start_end[0] = str(
                    dates[cell.column].date()) + ' ' + start_end[0]
                shift = {
                    'first_name': emps[cell.row].split(' ')[0],
                    'last_name': emps[cell.row].split(' ', 1)[1].replace(
                        " ", ""),
                    'shiftstart': start_end[0],
                    'shiftend': start_end[1],
                    'isic': IsIC,
                    'training': training,
                    'otbanked': OTBanked,
                    'otpaidout': OTPaidOut,
                    'comment': comment
                }

                return shift
        return None

    def create_start_end_time(self, value):
        start_end = []
        for time in value.split('-')[:2]:
            time = self.convert_to_24(time)
            start_end.append(time)
        return start_end

    def convert_to_24(self, time):
        time = time.split('\n')[0].replace(' ', '').strip()
        time, half_day = time[:-2], time[-2:].lower()
        if half_day == 'am':
            split = time.find(':')
            if split == -1:
                t = int(time)
                if t == 12:
                    t = '00'
                else:
                    t = str(t)
                return t + ':00'
            else:
                t = int(time[:split])
                if t == 12:
                    t = '00'
                else:
                    t = str(t)
                return t + time[split:]
        elif half_day == 'pm':
            split = time.find(':')
            if split == -1:
                t = int(time) + 12
                if t >= 24:
                    t = '00'
                else:
                    t = str(t)
                return t + ':00'
            else:
                t = int(time[:split]) + 12
                if t >= 24:
                    t = '00'
                else:
                    t = str(t)
                return t + time[split:]
        else:
            raise ValueError("Didn't finish with AM or PM." + time)


class TimeOff_Import(Staffhub_Import):
    timeoff_schema = {
        'first_name': {
            'type': 'string'
        },
        'last_name': {
            'type': 'string'
        },
        'dayof': {
            'type': 'date'
        },
        'partialtime': {
            'type': 'number'
        },
        'type': {
            'type': 'string',
            'allowed': ['Off', 'BDay', 'FRL', 'Sick', 'Vacation', 'Stat']
        },
        'comment': {
            'type': 'string',
            'nullable': True
        }
    }

    def __init__(self, excel_file, yearof):
        super().__init__(excel_file, yearof)
        self.data = self.generate_timeoff()
        self.valid_data = self.validate_import(self.data, self.timeoff_schema)

    def generate_timeoff(self):
        time_offs = []
        self.log['records'] = 0

        dates = self.generate_dates_dict(self.sh)
        emps = self.generate_employees_dict(self.sh)
        max_ = self.sh.max_column + 1

        for i in range(self.body_col_offset, max_):
            for ii in range(self.body_row_offset, self.sh.max_row):
                time_off = self.create_time_off_data(
                    self.sh.cell(ii, i), dates, emps)
                if time_off:
                    time_offs.append(time_off)
                    self.log['records'] = self.log['records'] + 1
                    print('added timeoff: ' + str(self.log['records']))
        self.log['type'] = 'time off'
        print('timeoffs generated: ' + str(self.log))
        return time_offs

    def create_time_off_data(self, cell, dates, emps):
        merge_value = self.check_merge(cell)
        if merge_value:
            m = merge_value.split('\n')
            comment = None
            if len(m) > 1:
                merge_value = m[0]
                comment = str(m[1])

            if merge_value in self.valid_time_off:
                time_off = self.create_time_off_record(
                    emps[cell.row], dates[cell.column].date(), merge_value,
                    comment)
                return time_off
        if cell.value:
            c = cell.value.split('\n')
            value = cell.value
            comment = None
            if len(c) > 1:
                value = c[0]
                comment = str(c[1])

            if value in self.valid_time_off:
                time_off = self.create_time_off_record(
                    emps[cell.row], dates[cell.column].date(), value, comment)
                return time_off
        return None

    def check_merge(self, cell):
        while True:
            if cell.coordinate in self.sh.merged_cells:
                if not cell.value:
                    cell = cell.offset(0, -1)
                elif cell.value:
                    return cell.value
            else:
                return None

    def create_time_off_record(self, employee, dayof, type_, comment):
        time_off = {
            'first_name': employee.split(' ')[0],
            'last_name': employee.split(' ', 1)[1].replace(" ", ""),
            'dayof': dayof,
            'partialtime': -1,
            'type': type_,
            'comment': comment
        }

        return time_off
